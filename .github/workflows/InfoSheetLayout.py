# -*- coding: utf-8 -*-
"""
Created on Wed Nov 27 08:59:33 2019

@author: User
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os
import re
from countdown import countdown

print('即将对集团和证券公司干部信息明细表进行格式美化... ')
seconds = 3
countdown(seconds)

date = '999999'
try:
    with open(r"E:\23-个人\month.txt",'r') as file_to_read:
        s = file_to_read.read()
    exec(s)
except:
    print('未知异常')

while re.match(r'\d{4}(1[0-2]{1}$|0[0-9]{1}$)', date) == None:
    date = input('输入的年月有误，请按格式重新输入6位年月，(格式：YYYYMM):')

print('请检查文件目录是否正确、确保目录下有以下文件：\n “E:\\1-统计\\%s\\raw\\干部信息明细表（数据清洗）.xlsx” ...' %date)
countdown(seconds)
if os.access(r'E:\1-统计\%s\raw\干部信息明细表（数据清洗）.xlsx' %date, os.F_OK):
    pass
else:
    input('【E:\\1-统计\\%s\\raw\\干部信息明细表（数据清洗）.xlsx】不存在，是否继续？（按回车键继续...） \n' %date)

wb = load_workbook(r'E:\1-统计\%s\raw\干部信息明细表（数据清洗）.xlsx' %date)

ws1 = wb['干部信息明细表总表']
ws2 = wb['集团和证券公司领导']
ws3 = wb['集团和证券总部干部']
ws4 = wb['分公司干部']
ws5 = wb['子公司干部']
ws6 = wb['营业部干部']
ws7 = wb['其他干部']
ws8 = wb['转岗人员']

#字体
font1 = Font(name='黑体', color='FFFFFF', size=10, b=True)
font2 = Font(name='宋体', color='000000', size=10)

#单元格格式
ali = Alignment(horizontal='center', vertical='center',wrap_text=True )

#颜色填充
fill1 = PatternFill('solid', fgColor='0070C0')
fill2 = PatternFill('solid', fgColor='FFFFFF')

#边线和边框
sd1 = Side(style='thin', color='000000')
sd2 = Side(style='medium', color='000000')
border1 = Border(top=sd1, bottom=sd2, left=sd1, right=sd1)
border2 = Border(top=sd1, bottom=sd1, left=sd1, right=sd1)

#打包以上格式
sty1 = NamedStyle(name='sty1', font=font1, fill=fill1,border=border1, alignment=ali)
sty2 = NamedStyle(name='sty2', font=font2, fill=fill2,border=border2, alignment=ali)

def setup(ws):
    ws.delete_cols(18, 14) #从18列开始删除，往后删14列（删除之前用pandas匹配的用于统计汇总的数据字段）

    rows = ws.max_row
    cols = ws.max_column

    #第一行表头和第二行开始的数据分别匹配格式
    for r in range(1, rows+1):
        for c in range(1, cols+1):
            if r == 1:
                ws.cell(r, c).style = sty1
                ws.row_dimensions[r].height = 27
            else:
                ws.cell(r, c).style = sty2

    #冻结首行
    ws.freeze_panes = 'A2'
    #设置列宽
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 27
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 9
    ws.column_dimensions["K"].width = 9
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 4
    ws.column_dimensions["O"].width = 4
    ws.column_dimensions["P"].width = 10
    ws.column_dimensions["Q"].width = 10

setup(ws1)
setup(ws2)
setup(ws3)
setup(ws4)
setup(ws5)
setup(ws6)
setup(ws7)
setup(ws8)

print('将输出文件至目录E:\\1-统计\\%s\\raw\\' %date)
countdown(seconds)
wb.save(r'E:\\1-统计\\%s\\raw\\集团和证券公司干部信息明细表%s.xlsx' %(date,date))
